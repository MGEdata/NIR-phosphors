import numpy as np
from sklearn.model_selection import GridSearchCV
from sklearn.kernel_ridge import KernelRidge
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler
from sklearn.utils import resample
import openpyxl
from scipy import stats
rng = np.random.RandomState(0)
np.set_printoptions(threshold=np.inf)
np.set_printoptions(suppress=True)


dataset = pd.read_csv("IQE_training set.csv")
x = dataset.iloc[:, 0:3]
y = dataset.iloc[:, 3]
ss_x = StandardScaler()
x = ss_x.fit_transform(x)
ss_y = StandardScaler()
y = ss_y.fit_transform(y.values.reshape(-1, 1))


def model(x, y, data_predict):
    param_grid = [{"alpha": [0.1, 0.15, 0.2, 0.05, 0.01], "gamma": np.logspace(-2, 2, 10)}]
    grid = GridSearchCV(KernelRidge(kernel='rbf'), param_grid, cv=10, scoring='neg_mean_absolute_error')
    grid.fit(x, y.ravel())
    rf1 = grid.best_estimator_
    model_y = rf1.fit(x, y.ravel())
    predictions = model_y.predict(data_predict)
    return(predictions)


def data_write(data, path):
    outwb = openpyxl.Workbook()
    ws = outwb.create_sheet(index=1)
    i = 1
    r = 1
    for line in data:
        for col in range(1, len(line) + 1):
            #ColNum = r
            ws.cell(row=r, column=col).value = line[col - 1]
        i += 1
        r += 1
    savexlsx = path
    outwb.save(savexlsx)


def PA(goal, searchspace0):
    p = (goal-searchspace0[:, 0])/(searchspace0[:, 1])
    pas = 1-stats.norm.cdf(p)
    return(pas)


dataset1 = pd.read_csv("IQE_To predict.csv")
X1 = dataset1.iloc[:, 0:3]
X22 = ss_x.transform(X1)
pre_a = np.arange(0, len(X22))
n = len(y)
m = 1000
for u in range(0, m):
    x11, y11 = resample(x, y, replace=True, n_samples=n, random_state=u)
    predict_data = model(x11, y11, X22)
    predict_y = ss_y.inverse_transform(predict_data)
    print(u)
    pre_a = np.column_stack((pre_a, predict_y))
pre_all = np.column_stack((np.mean(pre_a[:, 1:, ], axis=1), np.std(pre_a[:, 1:, ], axis=1)))
PA = PA(95, pre_all)
m01m = MinMaxScaler()
PAs = np.array(PA)
PA01 = m01m.fit_transform(PAs.reshape(-1, 1))
pre_all = np.column_stack((pre_all, PA))
pre_all = np.column_stack((pre_all, PA01))


data_write(pre_all, "102.xlsx")
